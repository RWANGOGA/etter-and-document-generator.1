import io
import json
import re
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import Response
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from groq import Groq

from app.config import settings

router = APIRouter(prefix="/api/xlsx", tags=["xlsx"])
client = Groq(api_key=settings.groq_api_key)


class ThemeOptions(BaseModel):
    primary_color: str = "1E3A5F"
    text_color: str = "000000"
    font_family: str = "Calibri"


class XlsxRequest(BaseModel):
    topic: str
    row_count: int = 15
    theme: ThemeOptions = ThemeOptions()


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]


class ChatResponse(BaseModel):
    success: bool
    reply: str
    structure: dict | None = None
    ready: bool = False


class BuildRequest(BaseModel):
    structure: dict
    theme: ThemeOptions = ThemeOptions()


def _strip_thinking(text: str) -> str:
    return re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()


def _extract_json(text: str) -> dict:
    text = _strip_thinking(text)
    text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip())
    start = text.find("{")
    end = text.rfind("}") + 1
    if start == -1 or end == 0:
        raise ValueError("No JSON object found in model output")
    return json.loads(text[start:end])


STRUCTURE_PROMPT = """You are generating structured data for an Excel spreadsheet.
Topic: {topic}
Target row count (excluding header): {row_count}

Return ONLY valid JSON, no markdown fences, no commentary, in exactly this shape:
{{
  "sheet_title": "Short title for the sheet tab",
  "columns": ["Column A", "Column B", "Column C"],
  "rows": [
    ["value1", "value2", "value3"]
  ]
}}

Rules:
- Choose 3-6 columns genuinely relevant to the topic.
- Produce exactly {row_count} data rows.
- Use realistic, specific values — not placeholder text like "Value 1".
- Keep each cell value short (under 40 characters)."""


def _generate_structure(topic: str, row_count: int) -> dict:
    prompt = STRUCTURE_PROMPT.format(topic=topic, row_count=row_count)
    completion = client.chat.completions.create(
        model="qwen/qwen3.6-27b",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=4096,
        temperature=0.6,
    )
    raw = completion.choices[0].message.content or ""
    return _extract_json(raw)


def _read_xlsx(content: bytes, max_rows: int = 300) -> dict:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return {"columns": [], "rows": []}

    columns = [str(c) if c is not None else "" for c in header]
    rows = []
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            break
        rows.append(["" if v is None else str(v) for v in row])

    return {"sheet_title": ws.title, "columns": columns, "rows": rows}


ENHANCE_PROMPT = """You are improving an existing spreadsheet.
Instructions from the user: {instructions}

Current sheet title: {sheet_title}
Current columns: {columns}
Current data (JSON rows, may be truncated):
{rows}

Return ONLY valid JSON, no markdown fences, no commentary, in exactly this shape:
{{
  "sheet_title": "Short title for the sheet tab",
  "columns": ["Column A", "Column B", "Column C"],
  "rows": [
    ["value1", "value2", "value3"]
  ]
}}

Rules:
- Apply the user's instructions: clean up messy data, fix inconsistent formatting, fill obviously missing values sensibly, restructure or add columns/rows if asked.
- Keep all genuinely useful original data — don't discard rows without reason.
- Keep each cell value short (under 40 characters)."""


def _enhance_structure(existing: dict, instructions: str) -> dict:
    prompt = ENHANCE_PROMPT.format(
        instructions=instructions,
        sheet_title=existing.get("sheet_title", "Sheet1"),
        columns=json.dumps(existing.get("columns", [])),
        rows=json.dumps(existing.get("rows", [])[:100]),
    )
    completion = client.chat.completions.create(
        model="qwen/qwen3.6-27b",
        messages=[{"role": "user", "content": prompt}],
        max_tokens=4096,
        temperature=0.5,
    )
    raw = completion.choices[0].message.content or ""
    return _extract_json(raw)


CHAT_SYSTEM_PROMPT = """You are a spreadsheet-building assistant. Have a short conversation with the user to understand exactly what spreadsheet they need: its purpose, what columns it should track, roughly how many rows, and any specific data points they already know.

Ask at most 2-3 clarifying questions total across the conversation — don't drag it out. Once you have enough to build something genuinely useful (or if the user says to just go ahead), respond with:
1. One short plain-text confirmation line (no markdown, no emojis).
2. Then the full structure as JSON inside <STRUCTURE>...</STRUCTURE> tags, in exactly this shape:
{
  "sheet_title": "Short title",
  "columns": ["Column A", "Column B"],
  "rows": [["value1", "value2"]]
}

Do not include <STRUCTURE> tags until you are actually ready to deliver the final structure — during clarifying questions, just talk normally in plain text, no markdown symbols."""


@router.post("/generate")
async def generate_xlsx(req: XlsxRequest):
    if not req.topic.strip():
        raise HTTPException(status_code=400, detail="Topic is required")
    try:
        structure = _generate_structure(req.topic, req.row_count)
        data = build_xlsx(structure, req.theme)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Spreadsheet generation failed: {e}")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=spreadsheet.xlsx"},
    )


@router.post("/enhance")
async def enhance_xlsx(
    file: UploadFile = File(...),
    instructions: str = Form(...),
    primary_color: str = Form("1E3A5F"),
    text_color: str = Form("000000"),
    font_family: str = Form("Calibri"),
):
    content = await file.read()
    try:
        existing = _read_xlsx(content)
        structure = _enhance_structure(existing, instructions)
        theme = ThemeOptions(primary_color=primary_color, text_color=text_color, font_family=font_family)
        data = build_xlsx(structure, theme)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Spreadsheet enhancement failed: {e}")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=enhanced.xlsx"},
    )


@router.post("/chat", response_model=ChatResponse)
async def chat_xlsx(req: ChatRequest):
    messages = [{"role": "system", "content": CHAT_SYSTEM_PROMPT}]
    for m in req.messages:
        messages.append({"role": m.role, "content": m.content})

    try:
        completion = client.chat.completions.create(
            model="qwen/qwen3.6-27b",
            messages=messages,
            max_tokens=4096,
            temperature=0.6,
        )
        text = _strip_thinking(completion.choices[0].message.content or "")

        structure = None
        reply_text = text
        ready = False

        if "<STRUCTURE>" in text and "</STRUCTURE>" in text:
            start = text.index("<STRUCTURE>")
            end = text.index("</STRUCTURE>") + len("</STRUCTURE>")
            raw_json = text[start + len("<STRUCTURE>"):text.index("</STRUCTURE>")].strip()
            reply_text = (text[:start] + text[end:]).strip()
            structure = json.loads(raw_json)
            ready = True
            if not reply_text:
                reply_text = "Here's the spreadsheet — review it and click Build to download."

        return ChatResponse(success=True, reply=reply_text, structure=structure, ready=ready)
    except Exception as e:
        return ChatResponse(success=False, reply=f"Something went wrong: {e}")


@router.post("/build")
async def build_from_structure(req: BuildRequest):
    try:
        data = build_xlsx(req.structure, req.theme)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Build failed: {e}")
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=spreadsheet.xlsx"},
    )


def build_xlsx(structure: dict, theme: ThemeOptions) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (structure.get("sheet_title") or "Sheet1")[:31]

    header_fill = PatternFill(start_color=theme.primary_color, end_color=theme.primary_color, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name=theme.font_family)
    body_font = Font(size=10, name=theme.font_family, color=theme.text_color)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    columns = structure.get("columns", [])
    rows = structure.get("rows", [])

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            cell.border = thin_border

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(r[col_idx - 1])) for r in rows if len(r) >= col_idx] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()